import os
from operator import attrgetter
from collections import defaultdict
from openpyxl import Workbook

from SemanticModels import SemanticEntity, SemanticProperty, SemanticData

def has_type(e, string):
    for p in e.outgoing:
        if p.short_type == "P2" and p.target.string == string: return True
    return False

if __name__ == "__main__":
    pickle_file = "../Data/ParsedSemanticAnnotations.pickle"
    save_filepath = "../Documentation/Visualizations/Wikidata_Locations.xlsx"
    
    data = SemanticData(pickle_file)
    
    
    locations = defaultdict(list)
    for e in data.entities:
        if e.short_type == "E53" and not has_type(e, "Building"):
            locations[e.search_string].append(e)
    
    wb = Workbook()

    # Cumulative
    ws = wb.active
    ws.title = "Locations"
    
    # header
    ws.append(("Normalized string", "Raw string", "Wikidata ID"))
    ws['A1'].style = "Headline 2"
    ws['B1'].style = "Headline 2"
    ws['C1'].style = "Headline 2"
    
    dimA, dimB = 1,1
    for string, places in locations.items():
        dimA = max(dimA, len(string))
        dimB = max(dimB, len( places[0].string))
        ws.append((string, places[0].string, ''))
        
    freeze = ws['A2']
    ws.freeze_panes = freeze
    ws.column_dimensions['A'].width = dimA +4
    ws.column_dimensions['B'].width = dimB +4
    ws.column_dimensions['C'].width = len("Wikidata ID") +4
    
    wb.save(filename=save_filepath)
    print(f"Saved {len(locations)} locations to {save_filepath}")