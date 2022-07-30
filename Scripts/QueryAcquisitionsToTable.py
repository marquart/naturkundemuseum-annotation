from cProfile import label
import os
from operator import attrgetter
from collections import defaultdict
import string
from openpyxl import Workbook

from SemanticModels import SemanticEntity, SemanticProperty, SemanticData

from ImportantAcquisitions import Acquisition, findAcquisitions


def append_header(ws, labels):
    assert len(labels)<27
    ws.append(labels)

    for letter in string.ascii_uppercase[:len(labels)]:
        ws[f'{letter}1'].style = "Headline 2"

    #ws.row_dimensions[1].style = "Title"
    #ws['A1:B1'].style = "Headline 1"

def stringifyNeighors(acq, types):
    sequence = ', '.join(e.string for e in acq.neighbors if not e.virtual and e.short_type in types)
    if sequence: return sequence
    return ''

if __name__ == "__main__":
    pickle_file = "../Data/ParsedSemanticAnnotations.pickle"
    save_filepath = "../Documentation/Visualizations/"
    
    data = SemanticData(pickle_file)
    
    acquisitions = findAcquisitions(data.entities)
    
    # Build Excel Table
    labels = ( "ID",
                "Year",
                "Page Number",
                "Line Number",
                "Acquisition Type",
                "Receiving Collection",
                "Giver(s)",
                "Object(s)",
                "Locations",
                "Taxonomic Term",
                "Dimension",
                "Condition Assessment",
                "Modification")

    wb = Workbook()
    
    # Cumulative
    ws = wb.active
    ws.title = "Acquisitions"
    append_header(ws, labels)
    
    dimensions = [len(l) for l in labels]

    for acq in acquisitions:
        row = ( acq.entity.id,
                acq.entity.year,
                acq.entity.page,
                acq.entity.line,
                acq.category,
                acq.holding.string if acq.holding else '',
                ', '.join(e.string for e in acq.givers) if acq.givers else '',
                stringifyNeighors(acq, ("E18","E19","E20")),
                ', '.join(e.string for e in acq.locations) if acq.locations else '',
                stringifyNeighors(acq, ("E28",)),
                stringifyNeighors(acq, ("E54",)),
                stringifyNeighors(acq, ("E14","E3")),
                stringifyNeighors(acq, ("E11",))
        )
        if len(row)!=len(labels):
            print(len(row))
            print(row)
        assert len(row) == len(labels)

        for i,entry in enumerate(row):
            if (new_len := len(str(entry))) > dimensions[i]: dimensions[i] = new_len

        ws.append(row)

    freeze = ws['A2']
    ws.freeze_panes = freeze

    for letter, dim in zip(string.ascii_uppercase, dimensions):
        ws.column_dimensions[letter].width = dim +2

    filepath = os.path.join(save_filepath, f"Chronik_Acquisitions.xlsx")
    wb.save(filename=filepath)
    print(f"Exported {len(acquisitions)} acquisitions to {filepath}")
    