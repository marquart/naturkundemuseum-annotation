import os
from operator import attrgetter
from collections import defaultdict
from openpyxl import Workbook

from SemanticModels import SemanticEntity, SemanticProperty, SemanticData

def has_type(e, string):
    for p in e.outgoing:
        if p.short_type == "P2" and p.target.string == string: return True
    return False

def append_header(ws, label):
    ws.append((label,"No. of Mentions", "Alternative Names"))
    ws['A1'].style = "Headline 2"
    ws['B1'].style = "Headline 2"
    ws['C1'].style = "Headline 2"
    #ws.row_dimensions[1].style = "Title"
    #ws['A1:B1'].style = "Headline 1"


def read_mentions(entities):
    return sum(e.mentions for e in entities)

def strlen(e):
    return len(e.string)

def optimal_person_name(entities):
    return min(entities, key=strlen).string

if __name__ == "__main__":
    pickle_file = "../Data/ParsedSemanticAnnotations.pickle"
    save_filepath = "../Documentation/Visualizations/"
    
    data = SemanticData(pickle_file)
    
    for t, label in (("E53","Locations"),("E21","Persons")):
        # Query Types
        cum_table  = defaultdict(list) # search_string : [...entities]
        year_table = defaultdict(dict) # year: search_string : [...entities]
        for e in data.entities:
            if e.short_type == t and not has_type(e, "Building"):
                cum_table[e.search_string].append(e)
                year = e.year
                if e.search_string in year_table[year]:
                    year_table[year][e.search_string].append(e)
                else:
                    year_table[year][e.search_string] = [e]
    
        # Build Excel Table
        wb = Workbook()
        
        # Cumulative
        ws = wb.active
        ws.title = "Cumulative"
        append_header(ws, label)
        
        dimA = 1
        dimB = len("No. of Mentions")
        dimC = 1
        
        for entities in sorted(cum_table.values(), key=read_mentions, reverse=True):
            if t == "E21": name = optimal_person_name(entities)
            else: name = entities[-1].string
            
            if len(name) > dimA: dimA = len(name)
            
            alternatives = '; '.join(set(e.string for e in entities if e.string != name))
            if len(alternatives) > 0:
                if len(alternatives) > dimC: dimC = len(alternatives)
            else:
                alternatives = ""
            
            ws.append((name, read_mentions(entities), alternatives))
        freeze = ws['A2']
        ws.freeze_panes = freeze
        ws.column_dimensions['A'].width = dimA +4
        ws.column_dimensions['B'].width = dimB +4
        ws.column_dimensions['C'].width = dimC +4
        
        # per year
        for year in sorted(year_table):
            ws = wb.create_sheet(title=str(year))
            append_header(ws, label)
            
            entities_lst = year_table[year].values()
            
            for entities in sorted(entities_lst, key=read_mentions, reverse=True):
                if t == "E21": name = optimal_person_name(entities)
                else: name = entities[-1].string
                
                alternatives = '; '.join(set(e.string for e in entities if e.string != name))
                if len(alternatives) > 0:
                    if len(alternatives) > dimC: dimC = len(alternatives)
                else:
                    alternatives = ""
                
                ws.append((name, read_mentions(entities), alternatives))
                
            ws.column_dimensions['A'].width = dimA +4
            ws.column_dimensions['B'].width = dimB +4
            ws.column_dimensions['C'].width = dimC +4
            
            freeze = ws['A2']
            ws.freeze_panes = freeze
        filepath = os.path.join(save_filepath, f"Chronik_{label}.xlsx")
        wb.save(filename=filepath)
        print(f"Calculated mentions for {label} and saved to {filepath}")
    