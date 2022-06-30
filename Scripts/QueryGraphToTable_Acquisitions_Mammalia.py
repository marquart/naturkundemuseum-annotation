import os
from operator import attrgetter
from collections import defaultdict
from openpyxl import Workbook

from SemanticModels import SemanticEntity, SemanticProperty, SemanticData

def has_type(e, string):
    for p in e.outgoing:
        if p.short_type == "P2" and p.target.string == string: return True
    return False


def find_places(entities, placeName):
    results = set()
    for e in entities:
        if e.short_type == "E53":
            cursor = e
            while cursor is not None:
                if placeName in cursor.search_string:
                    results.add(e)
                    print(f"    Included Location: {str(e)}")
                    break
                next_cursor = None
                for p in cursor.outgoing:
                    if p.short_type == "P89":
                        next_cursor = p.target
                        break
                if next_cursor is None: break
                else: cursor = next_cursor
    return results

def find_holding(entities, collectionName):
    results = set()
    for e in entities:
        if e.short_type == "E78":
            cursor = e
            while cursor is not None:
                if collectionName in cursor.search_string:
                    results.add(e)
                    print(f"    Included Collection: {str(e)}")
                    break
                next_cursor = None
                for p in cursor.incoming:
                    if p.short_type == "P46":
                        next_cursor = p.source
                        break
                if next_cursor is None: break
                else: cursor = next_cursor
    return results

def acquisitionsConnectedToLocation(collection, locations):
    #assert collection.short_type in ("E8","E96") and locations[0].short_type == "E53"
    
    acquisitionsConnectedToLocations = set()
    for p in collection.incoming:
        if p.short_type == "P22":
            for pp in p.source.outgoing: # Acquisition
                if pp.short_type == "P24":
                    for ppp in pp.target.outgoing: # Object
                        if ppp.short_type == "P53" and ppp.target in locations:
                            acquisitionsConnectedToLocations.add(p.source)
                            break
                        elif ppp.short_type == "P46":
                            for pppp in ppp.target.outgoing: # Object
                                if pppp.short_type == "P53" and pppp.target in locations:
                                    acquisitionsConnectedToLocations.add(p.source)
                                    break
                elif pp.short_type == "P23":
                    for ppp in pp.target.outgoing: # Person
                        if ppp.short_type == "P53" and ppp.target in locations:
                            acquisitionsConnectedToLocations.add(p.source)
    return acquisitionsConnectedToLocations


def get_collectors(acquisitions):
    for e in acquisitions:
        found_collector = False
        for p in e.outgoing:
            if p.short_type == "P23":
                found_collector = True
                yield p.target, e
        if not found_collector:
            yield None, e


def sum_by_acquisition_type(acquisitions):
    result = [0,0,0]
    for e in acquisitions:
        if e.short_type == "E96": result[1] += 1
        elif e.short_type == "E8":
            if has_type(e, "Trade"): result[2] += 1
            else: result[0] += 1
    return result

def strlen(e):
    return len(e.string)

def optimal_person_name(entities):
    return min(entities, key=strlen).string

def distinct_person_name(e):
    return e.string.split(' ')[-1].lower()


def append_header(ws):
    ws.append(("Collector Name", "Donations", "Purchases",  "Trades", "Alternative Collector Names"))
    ws['A1'].style = "Headline 2"
    ws['B1'].style = "Headline 2"
    ws['C1'].style = "Headline 2"
    ws['D1'].style = "Headline 2"
    ws['E1'].style = "Headline 2"

    

def build_acquisition_types_table(entities, save_filepath, filter_collections=('',), filter_locations=('',)):
    filtered_collections = find_holding(data.entities, filter_collections[0])
    for filter in filter_collections[1:]:
        filtered_collections.update(find_holding(data.entities, filter))
    
    
    filtered_locations = find_places(data.entities, filter_locations[0])
    for filter in filter_locations[1:]:
        filtered_locations.update(find_places(data.entities, filter))
    #print('\n'.join(e.string for e in kamerun))
    
    personName_entities = defaultdict(list)
    cum_table  = defaultdict(list) # distinct_person_name : [...acquisitions]
    year_table = defaultdict(dict) # year: distinct_person_name : [...acquisitions]
    for collection in filtered_collections:
        acquisitions = acquisitionsConnectedToLocation(collection, filtered_locations)
        
        for actor, acquisition in get_collectors(acquisitions):
            if actor:
                name = distinct_person_name(actor)
                personName_entities[name].append(actor)
            else:
                name = "-"
            cum_table[name].append(acquisition)
            if name in year_table[acquisition.year]: year_table[acquisition.year][name].append(acquisition)
            else: year_table[acquisition.year][name] = [acquisition]
    
    # Build Excel
    wb = Workbook()

    # Cumulative
    ws = wb.active
    ws.title = "Total"
    append_header(ws)

    dimA = 1
    dimB = 10
    dimC = 1
    
    for identificator in sorted(cum_table.keys(), key=lambda x: len(cum_table[x]), reverse=True):
        acquisitions = cum_table[identificator]
        
        if (collectors := personName_entities[identificator]):
            name = optimal_person_name(collectors)
        else:
            name = '-'
        
        if len(name) > dimA: dimA = len(name)
        
        alternatives = '; '.join(set(e.string for e in personName_entities[identificator] if e.string != name))
        if len(alternatives) > 0:
            if len(alternatives) > dimC: dimC = len(alternatives)
        else:
            alternatives = ""
        
        counts = sum_by_acquisition_type(acquisitions)
        
        ws.append((name, counts[0],counts[1],counts[2], alternatives))
        
    freeze = ws['A2']
    ws.freeze_panes = freeze
    ws.column_dimensions['A'].width = dimA +4
    ws.column_dimensions['B'].width = dimB +4
    ws.column_dimensions['C'].width = dimB +4
    ws.column_dimensions['D'].width = dimB +4
    ws.column_dimensions['E'].width = dimC +4
    
    # per year
    for year in sorted(year_table):
        ws = wb.create_sheet(title=str(year))
        append_header(ws)
        
        names = year_table[year].keys()
        
        for identificator in sorted(names, key=lambda x: len(year_table[year][x]), reverse=True):
            acquisitions = year_table[year][identificator]
            
            if (collectors := personName_entities[identificator]):
                name = optimal_person_name(collectors)
            else:
                name = '-'
            
            if len(name) > dimA: dimA = len(name)
            
            alternatives = '; '.join(set(e.string for e in personName_entities[identificator] if e.string != name))
            if len(alternatives) > 0:
                if len(alternatives) > dimC: dimC = len(alternatives)
            else:
                alternatives = ""
            
            counts = sum_by_acquisition_type(acquisitions)
            
            ws.append((name, counts[0],counts[1],counts[2], alternatives))
            
        ws.column_dimensions['A'].width = dimA +4
        ws.column_dimensions['B'].width = dimB +4
        ws.column_dimensions['C'].width = dimB +4
        ws.column_dimensions['D'].width = dimB +4
        ws.column_dimensions['E'].width = dimC +4
        
        freeze = ws['A2']
        ws.freeze_panes = freeze
    
    #filepath = os.path.join(save_filepath, f"Chronik_Mammalia-Collectors_Kamerun.xlsx")
    wb.save(filename=save_filepath)
    print(f"Saved to {save_filepath}")
    
    
if __name__ == "__main__":
    pickle_file = "../Data/ParsedSemanticAnnotations.pickle"
    save_filepath = "../Documentation/Visualizations/Chronik_Mammalia-Collectors_Kamerun.xlsx"
    
    data = SemanticData(pickle_file)
    
    build_acquisition_types_table(data.entities, save_filepath, filter_collections=("ammal", "äuget"), filter_locations=('amerun',))
    
    save_filepath = "../Documentation/Visualizations/Chronik_Collectors_Kamerun.xlsx"
    build_acquisition_types_table(data.entities, save_filepath, filter_collections=("",), filter_locations=('amerun',))

    save_filepath = "../Documentation/Visualizations/Chronik_Collectors_Australien.xlsx"
    build_acquisition_types_table(data.entities, save_filepath, filter_collections=("",), filter_locations=('strali',))
